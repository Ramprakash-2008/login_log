from flask import Flask, request, redirect, url_for, render_template_string, session, send_file
import sqlite3, os, smtplib
from datetime import datetime
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime, time
import pytz
from dotenv import load_dotenv
load_dotenv()
app = Flask(__name__)
app.secret_key = "super_secret_key"
 
# === CONFIG ===
ADMIN_PASSWORD = os.getenv("ADMIN_PASS")
ADMIN_EMAIL = os.getenv("EMAIL_USER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASS")
          # <-- Replace this
DATABASE = "users.db"
LOG_FILE = "login_log.csv"
DEADLINE_TIME = "09:00"

# === HTML Templates ===
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang='en'>
<head>
    <meta charset='UTF-8'>
    <meta name='viewport' content='width=device-width, initial-scale=1.0'>
    <title>Admin Login</title>
    <style>
        body {
            background-color: #fafafa;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            font-family: 'Segoe UI', sans-serif;
        }
        .login-box {
            background: white;
            border: 1px solid #dbdbdb;
            padding: 40px;
            width: 350px;
            text-align: center;
        }
        h2 {
            font-size: 30px;
            margin-bottom: 20px;
        }
        input[type="password"] {
            width: 100%;
            padding: 10px;
            margin-bottom: 10px;
            border: 1px solid #ccc;
            border-radius: 4px;
        }
        button {
            width: 100%;
            padding: 10px;
            background-color: #3897f0;
            color: white;
            border: none;
            border-radius: 4px;
            font-weight: bold;
            cursor: pointer;
        }
        a {
            display: block;
            margin-top: 15px;
            font-size: 14px;
            color: #00376b;
            text-decoration: none;
        }
    </style>
</head>
<body>
    <form method="POST">
        <div class="login-box">
            <h2>Admin</h2>
            <input type="password" name="password" placeholder="Password" required>
            <button type="submit">Log In</button>
            <a href="/">← Back to Login</a>
        </div>
    </form>
</body>
</html>

"""

ADMIN_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Admin Dashboard</title>
    <style>
        body { font-family: Arial; background: #f4f4f4; margin: 0; padding: 20px; }
        .container { max-width: 1000px; margin: auto; background: #fff; padding: 20px; border-radius: 8px; }
        .topbar button { margin-right: 10px; }
        .hidden { display: none; }
        .status { color: green; font-weight: bold; }
        .late { color: red; }
        .ontime { color: green; }
        table { width: 100%; border-collapse: collapse; }
        table, th, td { border: 1px solid #ddd; }
        th, td { padding: 10px; text-align: center; }
        button { padding: 10px; margin-top: 10px; border-radius: 5px; border: none; background: #3897f0; color: #fff; cursor: pointer; }
        button:hover { background: #287bd1; }
        h2, h3 { text-align: center; }
    </style>
    <script>
        function toggle(id) {
            const el = document.getElementById(id);
            el.style.display = (el.style.display === "none") ? "block" : "none";
        }
    </script>
</head>
<body>
<div class="container">
<h2>Admin Dashboard - Login Logs</h2>
<div class="topbar">
    <button onclick="toggle('addUserForm')">➕ Add User</button>
    <button onclick="toggle('updateUsersForm')">🛠 Update Users</button>
    <form method="GET" action="/users" style="display:inline;">
        <button>👤 Show All Registered Users</button>
    </form>
    <form method="GET" action="/download-log" style="display:inline;">
        <button>📥 Download Excel Log</button>
    </form>
    <form method="GET" action="/logout" style="display:inline;">
        <button>🔒 Logout</button>
    </form>
</div>

{% if message %}<p class="status">✅ {{ message }}</p>{% endif %}

<div id="addUserForm" class="hidden">
    <form method="POST">
        <label>Add Single Username:</label>
        <input type="text" name="new_username" required>
        <button type="submit">Add</button>
    </form>
    <br><hr>
    <form method="POST" enctype="multipart/form-data">
        <label>Upload Usernames File (.txt/.csv):</label>
        <input type="file" name="file" accept=".txt,.csv" required>
        <button type="submit">Upload</button>
    </form>
</div>

<div id="updateUsersForm" class="hidden">
    <form method="POST">
        <table>
            <tr><th>Username</th><th>Edit</th><th>Delete</th></tr>
            {% for user in users %}
            <tr>
                <td><input type="text" name="usernames" value="{{ user }}"></td>
                <td><input type="checkbox" name="edit_{{ loop.index0 }}"></td>
                <td><input type="checkbox" name="delete_{{ loop.index0 }}"></td>
                <input type="hidden" name="original_{{ loop.index0 }}" value="{{ user }}">
            </tr>
            {% endfor %}
        </table>
        <input type="hidden" name="count" value="{{ users|length }}">
        <button type="submit" name="save_changes">💾 Save Changes</button>
    </form>
</div>

<h3>Login Logs</h3>
<table>
<tr><th>Username</th><th>Time</th><th>Status</th></tr>
{% for row in logs %}
<tr>
    <td>{{ row[0] }}</td>
    <td>{{ row[1] }}</td>
    <td class="{{ 'late' if row[2] == 'Late' else 'ontime' }}">{{ row[2] }}</td>
</tr>
{% endfor %}
</table>
</div>
</body>
</html>

"""

USERS_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>All Registered Users</title>
    <style>
        body { font-family: Arial; background: #fafafa; padding: 20px; text-align: center; }
        ul { list-style-type: none; padding: 0; }
        li { padding: 5px; font-size: 18px; }
        a { text-decoration: none; color: #3897f0; font-weight: bold; }
    </style>
</head>
<body>
    <h2>All Registered Users</h2>
    <ul>
    {% for user in users %}
        <li>{{ user }}</li>
    {% endfor %}
    </ul>
    <a href="/admin">⬅ Back to Admin</a>
</body>
</html>

"""

# === Helper Functions ===
def init_db():
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS users (username TEXT UNIQUE)")
    conn.commit()
    conn.close()
init_db()
def get_users():
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute("SELECT username FROM users")
    users = [row[0] for row in cur.fetchall()]
    conn.close()
    return users

def save_log(username, time, status):
    if not os.path.exists(LOG_FILE):
        with open(LOG_FILE, 'w') as f:
            f.write("Username,Time,Status\n")
    with open(LOG_FILE, 'a') as f:
        f.write(f"{username},{time},{status}\n")

def get_logs():
    if not os.path.exists(LOG_FILE): return []
    with open(LOG_FILE) as f:
        next(f)
        return [line.strip().split(',') for line in f]

def send_late_email(username, login_time):
    msg = MIMEText(f"User '{username}' logged in late at {login_time}.")
    msg['Subject'] = f"LATE LOGIN: {username}"
    msg['From'] = ADMIN_EMAIL
    msg['To'] = ADMIN_EMAIL
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(ADMIN_EMAIL, EMAIL_PASSWORD)
            server.sendmail(ADMIN_EMAIL, ADMIN_EMAIL, msg.as_string())
    except Exception as e:
        print("❌ Email failed:", e)

@app.route("/download-log")
def download_excel():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    if not os.path.exists(LOG_FILE):
        return "Log file not found."

    wb = Workbook()
    ws = wb.active
    ws.title = "Login Log"
    ws.append(["Username", "Time", "Status"])

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    with open(LOG_FILE, "r") as f:
        next(f)
        for line in f:
            username, time, status = line.strip().split(",")
            row = [username, time, status]
            ws.append(row)
            fill = green if status == "On-time" else red
            for cell in ws[ws.max_row]:
                cell.fill = fill

    file_path = "login_log.xlsx"
    wb.save(file_path)
    return send_file(file_path, as_attachment=True)

@app.route("/")
def home():
    return render_template_string('''
    <!DOCTYPE html>
    <html lang='en'>
    <head>
        <meta charset='UTF-8'>
        <meta name='viewport' content='width=device-width, initial-scale=1.0'>
        <title>User Login</title>
        <style>
            body {
                background-color: #fafafa;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                font-family: 'Segoe UI', sans-serif;
            }
            .login-box {
                background: white;
                border: 1px solid #dbdbdb;
                padding: 40px;
                width: 350px;
                text-align: center;
            }
            h2 {
                font-size: 28px;
                margin-bottom: 20px;
            }
            input[type="text"] {
                width: 100%;
                padding: 10px;
                margin-bottom: 10px;
                border: 1px solid #ccc;
                border-radius: 4px;
            }
            button {
                width: 100%;
                padding: 10px;
                background-color: #3897f0;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                cursor: pointer;
            }
            a {
                display: block;
                margin-top: 15px;
                font-size: 14px;
                color: #00376b;
                text-decoration: none;
            }
        </style>
    </head>
    <body>
        <form method="POST" action="/login">
            <div class="login-box">
                <h2>User Login</h2>
                <input type="text" name="username" placeholder="Enter username" required>
                <button type="submit">Log In</button>
                <a href="/admin/login">🔐 Admin Login</a>
            </div>
        </form>
    </body>
    </html>
    ''')


@app.route("/login", methods=["POST"])
def login_user():
    try:
        username = request.form.get("username", "").strip()
        if not username:
            return "❌ Username required", 400

        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        c.execute("SELECT username FROM users WHERE username=?", (username,))
        result = c.fetchone()
        conn.close()

        if not result:
            return render_template_string("""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Login Failed</title>
                <style>
                    body {
                        background-color: #fafafa;
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        height: 100vh;
                        font-family: 'Segoe UI', sans-serif;
                    }
                    .login-box {
                        background: white;
                        border: 1px solid #dbdbdb;
                        padding: 40px;
                        width: 350px;
                        text-align: center;
                    }
                    h2 {
                        color: red;
                        margin-bottom: 20px;
                    }
                    a {
                        display: block;
                        margin-top: 15px;
                        font-size: 14px;
                        color: #00376b;
                        text-decoration: none;
                    }
                </style>
            </head>
            <body>
                <div class="login-box">
                    <h2>❌ You are not a registered user</h2>
                    <a href="/">← Back to Login</a>
                </div>
            </body>
            </html>
            """)

        india = pytz.timezone("Asia/Kolkata")
        now = datetime.now(india)
        current_time = now.time()
        deadline = time(hour=9, minute=0)

        status = "Late" if current_time > deadline else "On-time"
        log_time = now.strftime("%H:%M")

        save_log(username, log_time, status)

        if status == "Late":
            send_late_email(username, log_time)

        # Success page
        return render_template_string(f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Login Success</title>
            <style>
                body {{
                    background-color: #fafafa;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100vh;
                    font-family: 'Segoe UI', sans-serif;
                }}
                .login-box {{
                    background: white;
                    border: 1px solid #dbdbdb;
                    padding: 40px;
                    width: 350px;
                    text-align: center;
                }}
                h2 {{
                    color: green;
                    margin-bottom: 20px;
                }}
                p {{
                    font-size: 16px;
                }}
                a {{
                    display: block;
                    margin-top: 15px;
                    font-size: 14px;
                    color: #00376b;
                    text-decoration: none;
                }}
            </style>
        </head>
        <body>
            <div class="login-box">
                <h2>✅ Login Successful</h2>
                <p><strong>{username}</strong> logged in at <strong>{log_time}</strong></p>
                <p>Status: <span style="color:{'red' if status == 'Late' else 'green'};"><strong>{status}</strong></span></p>
                <a href="/">← Back to Login</a>
            </div>
        </body>
        </html>
        """)

    except Exception as e:
        print("❌ ERROR in /login:", e)
        return "Internal Server Error", 500
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    message = ""

    if request.method == "POST":
        if "new_username" in request.form:
            username = request.form["new_username"].strip()
            if username:
                conn = sqlite3.connect(DATABASE)
                cur = conn.cursor()
                cur.execute("INSERT OR IGNORE INTO users (username) VALUES (?)", (username,))
                conn.commit()
                conn.close()
                message = f"User '{username}' added."

        elif "file" in request.files:
            file = request.files["file"]
            if file:
                lines = file.read().decode("utf-8").splitlines()
                conn = sqlite3.connect(DATABASE)
                cur = conn.cursor()
                for name in lines:
                    cur.execute("INSERT OR IGNORE INTO users (username) VALUES (?)", (name.strip(),))
                conn.commit()
                conn.close()
                message = f"{len(lines)} users uploaded."

        elif "save_changes" in request.form:
            count = int(request.form["count"])
            conn = sqlite3.connect(DATABASE)
            cur = conn.cursor()
            for i in range(count):
                original = request.form[f"original_{i}"].strip()
                new_value = request.form.getlist("usernames")[i].strip()
                delete = request.form.get(f"delete_{i}")

                if delete:
                    cur.execute("DELETE FROM users WHERE username = ?", (original,))
                elif new_value != original:
                    cur.execute("UPDATE users SET username = ? WHERE username = ?", (new_value, original))
            conn.commit()
            conn.close()
            message = "Changes saved successfully."

    return render_template_string(ADMIN_TEMPLATE, logs=get_logs(), users=get_users(), message=message)

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form["password"] == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect("/admin")
        else:
            return render_template_string("""
                <h2 style=\"color:red;\">❌ Incorrect Password</h2>
                <a href=\"/admin/login\">🔁 Try Again</a>
            """)
    return render_template_string(LOGIN_TEMPLATE)

@app.route("/users")
def show_users():
    return render_template_string(USERS_TEMPLATE, users=get_users())

@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/admin/login")

# === Main ===
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
# === Helper Functions ===
def init_db():
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS users (username TEXT UNIQUE)")
    conn.commit()
    conn.close()
init_db()
def get_users():
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute("SELECT username FROM users")
    users = [row[0] for row in cur.fetchall()]
    conn.close()
    return users

def save_log(username, time, status):
    if not os.path.exists(LOG_FILE):
        with open(LOG_FILE, 'w') as f:
            f.write("Username,Time,Status\n")
    with open(LOG_FILE, 'a') as f:
        f.write(f"{username},{time},{status}\n")

def get_logs():
    if not os.path.exists(LOG_FILE): return []
    with open(LOG_FILE) as f:
        next(f)
        return [line.strip().split(',') for line in f]

def send_late_email(username, login_time):
    msg = MIMEText(f"User '{username}' logged in late at {login_time}.")
    msg['Subject'] = f"LATE LOGIN: {username}"
    msg['From'] = ADMIN_EMAIL
    msg['To'] = ADMIN_EMAIL
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(ADMIN_EMAIL, EMAIL_PASSWORD)
            server.sendmail(ADMIN_EMAIL, ADMIN_EMAIL, msg.as_string())
    except Exception as e:
        print("❌ Email failed:", e)

@app.route("/download-log")
def download_excel():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    if not os.path.exists(LOG_FILE):
        return "Log file not found."

    wb = Workbook()
    ws = wb.active
    ws.title = "Login Log"
    ws.append(["Username", "Time", "Status"])

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    with open(LOG_FILE, "r") as f:
        next(f)
        for line in f:
            username, time, status = line.strip().split(",")
            row = [username, time, status]
            ws.append(row)
            fill = green if status == "On-time" else red
            for cell in ws[ws.max_row]:
                cell.fill = fill

    file_path = "login_log.xlsx"
    wb.save(file_path)
    return send_file(file_path, as_attachment=True)

@app.route("/")
def home():
    return """
    <h2>User Login</h2>
    <form method="POST" action="/login">
        Username: <input type="text" name="username">
        <button type="submit">Login</button>
    </form>
    <a href="/admin/login">🔐 Admin Login</a>
    """

@app.route("/login", methods=["POST"])
def login_user():
    try:
        username = request.form.get("username", "").strip()
        if not username:
            return "❌ Username required", 400

        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        c.execute("SELECT username FROM users WHERE username=?", (username,))
        result = c.fetchone()
        conn.close()

        if not result:
            return render_template_string("""
            <h2 style=\"color:red;\">❌ You are not a registered user</h2>
            <a href=\"/\" style=\"text-decoration:none; font-size:18px;\">🔙 Back to Login</a>
            """)

        india = pytz.timezone("Asia/Kolkata")
        now = datetime.now(india)
        current_time = now.time()
        deadline = time(hour=9, minute=0)

        status = "Late" if current_time > deadline else "On-time"
        log_time = now.strftime("%H:%M")

        save_log(username, log_time, status)

        if status == "Late":
            send_late_email(username, log_time)

        return f"✅ {username} logged in at {log_time} ({status})"

    except Exception as e:
        print("❌ ERROR in /login:", e)
        return "Internal Server Error", 500

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    message = ""

    if request.method == "POST":
        if "new_username" in request.form:
            username = request.form["new_username"].strip()
            if username:
                conn = sqlite3.connect(DATABASE)
                cur = conn.cursor()
                cur.execute("INSERT OR IGNORE INTO users (username) VALUES (?)", (username,))
                conn.commit()
                conn.close()
                message = f"User '{username}' added."

        elif "file" in request.files:
            file = request.files["file"]
            if file:
                lines = file.read().decode("utf-8").splitlines()
                conn = sqlite3.connect(DATABASE)
                cur = conn.cursor()
                for name in lines:
                    cur.execute("INSERT OR IGNORE INTO users (username) VALUES (?)", (name.strip(),))
                conn.commit()
                conn.close()
                message = f"{len(lines)} users uploaded."

        elif "save_changes" in request.form:
            count = int(request.form["count"])
            conn = sqlite3.connect(DATABASE)
            cur = conn.cursor()
            for i in range(count):
                original = request.form[f"original_{i}"].strip()
                new_value = request.form.getlist("usernames")[i].strip()
                delete = request.form.get(f"delete_{i}")

                if delete:
                    cur.execute("DELETE FROM users WHERE username = ?", (original,))
                elif new_value != original:
                    cur.execute("UPDATE users SET username = ? WHERE username = ?", (new_value, original))
            conn.commit()
            conn.close()
            message = "Changes saved successfully."

    return render_template_string(ADMIN_TEMPLATE, logs=get_logs(), users=get_users(), message=message)

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form["password"] == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect("/admin")
        else:
            return render_template_string("""
                <h2 style=\"color:red;\">❌ Incorrect Password</h2>
                <a href=\"/admin/login\">🔁 Try Again</a>
            """)
    return render_template_string(LOGIN_TEMPLATE)

@app.route("/users")
def show_users():
    return render_template_string(USERS_TEMPLATE, users=get_users())

@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/admin/login")

# === Main ===
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
